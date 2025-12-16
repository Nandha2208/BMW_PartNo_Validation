import pandas as pd
import re
import os
import datetime
import streamlit as st
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ------------------ UTILITY ------------------
def clean_part(part):
    if pd.isna(part):
        return ""
    part = str(part).strip()
    part = re.sub(r"^\(.*?\)", "", part)
    part = re.split(r"[-_]", part)[0]
    part = re.sub(r"[^\w]", "", part)
    return part.upper()

# ------------------ MERGE 16 BOMs ------------------
def merge_16_boms(uploaded_files):
    all_bom_data = []

    for file in uploaded_files:
        df = pd.read_excel(file)
        base_name = os.path.splitext(file.name)[0]
        bom_code = base_name.rsplit('_', 1)[0]

        df.insert(0, 'Source File', bom_code)

        level_col = next((c for c in df.columns if c.lower() in ['dglvl','dg_lvl','bom level']), None)
        desc_col = next((c for c in df.columns if c.lower() in ['ojtxp','description']), None)

        if not level_col or not desc_col:
            continue

        df[level_col] = df[level_col].astype(str).str.strip()
        df['GA Name'] = None
        df.loc[df[level_col] == '.1', 'GA Name'] = df[desc_col]
        df['GA Name'] = df['GA Name'].ffill()

        ga = df.pop('GA Name')
        df.insert(1, 'GA Name', ga)

        all_bom_data.append(df)

    combined_df = pd.concat(all_bom_data, ignore_index=True)
    return combined_df

# ------------------ BMW VALIDATION ------------------
def run_bmw_validation(bmw_file, tvs_df):
    df_bmw = pd.read_excel(bmw_file, sheet_name="Structure Report", header=2, dtype=str)
    df_bmw.columns = df_bmw.columns.str.strip()
    df_bmw = df_bmw[["Material", "3rd Party Company No.", "part version (AI)"]]
    df_bmw.columns = ["BMW_Part", "TVS_Part", "BMW_AI"]
    df_bmw = df_bmw.dropna(subset=["TVS_Part"])
    df_bmw["TVS_Part"] = df_bmw["TVS_Part"].apply(clean_part)
    df_bmw = df_bmw.drop_duplicates(subset="TVS_Part")

    df_tvs = tvs_df.copy()
    df_tvs.columns = df_tvs.columns.str.strip()

    required_cols = ["Component", "BMW part no.", "Part No. needed at BMW", "BMW AI index", "Description"]
    df_tvs = df_tvs[[c for c in required_cols if c in df_tvs.columns]]
    df_tvs.columns = ["TVS_Part", "BMW_Part_TVS", "Needed at BMW", "TVS_AI", "Description"]
    df_tvs = df_tvs.dropna(subset=["TVS_Part"])
    df_tvs["TVS_Part"] = df_tvs["TVS_Part"].apply(clean_part)
    df_tvs = df_tvs.drop_duplicates(subset="TVS_Part")

    merged = pd.merge(df_tvs, df_bmw, on="TVS_Part", how="left")

    merged["Validation_Status"] = merged.apply(
        lambda r: "Missing BMW" if str(r["Needed at BMW"]).upper()=="YES" and pd.isna(r["BMW_Part_TVS"]) else "", axis=1)

    merged["Comparison"] = merged.apply(
        lambda r: "Mismatch" if pd.notna(r["BMW_Part_TVS"]) and pd.notna(r["BMW_Part"]) and str(r["BMW_Part_TVS"])!=str(r["BMW_Part"]) else "", axis=1)

    merged["AI_Comparison"] = merged.apply(
        lambda r: "Mismatch" if pd.notna(r["TVS_AI"]) and pd.notna(r["BMW_AI"]) and str(r["TVS_AI"])!=str(r["BMW_AI"]) else "", axis=1)

    return merged

# ------------------ STREAMLIT UI ------------------
st.set_page_config(page_title="BMW BOM Automation", layout="wide")

st.title("🚀 BMW Part No Validation – End-to-End Automation")

st.markdown("""
**This tool automates:**
- 16 BOM Merge  
- BMW Part No Validation  
- AI Index Validation  
- Final Report Generation  
""")

uploaded_boms = st.file_uploader(
    "Upload 16 TVS BOM Excel Files",
    type=["xlsx","xls"],
    accept_multiple_files=True
)

bmw_file = st.file_uploader(
    "Upload BMW BOM (Structure Report)",
    type=["xlsx","xlsm"]
)

if st.button("▶ Run Full Automation"):
    if not uploaded_boms or not bmw_file:
        st.warning("Please upload all required files")
    else:
        with st.spinner("Merging BOMs..."):
            merged_bom = merge_16_boms(uploaded_boms)

        with st.spinner("Running BMW Validation..."):
            result_df = run_bmw_validation(bmw_file, merged_bom)

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"BMW_Validation_Result_{timestamp}.xlsx"
        result_df.to_excel(output_file, index=False)

        st.success("✅ Automation Completed Successfully")

        with open(output_file, "rb") as f:
            st.download_button(
                label="📥 Download Result Excel",
                data=f,
                file_name=output_file,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
